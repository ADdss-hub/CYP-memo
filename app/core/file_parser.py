"""文件解析器模块"""
import os
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO


def get_file_type(file_path: str) -> str:
    """获取文件类型
    
    Args:
        file_path: 文件路径
    
    Returns:
        文件类型
    """
    # 使用文件扩展名判断文件类型，替代libmagic
    ext = os.path.splitext(file_path)[1].lower()
    mime_types = {
        '.pdf': 'application/pdf',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.doc': 'application/msword',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.gif': 'image/gif',
        '.bmp': 'image/bmp',
        '.txt': 'text/plain',
        '.md': 'text/markdown',
        '.html': 'text/html'
    }
    return mime_types.get(ext, 'application/octet-stream')


def parse_pdf(file_path: str) -> str:
    """解析PDF文件
    
    Args:
        file_path: PDF文件路径
    
    Returns:
        解析后的文本内容
    """
    content = []
    try:
        with open(file_path, 'rb') as f:
            reader = PdfReader(f)
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    content.append(text)
    except Exception as e:
        raise Exception(f"解析PDF文件失败: {str(e)}")
    
    return '\n'.join(content)


def parse_docx(file_path: str) -> str:
    """解析Word文档
    
    Args:
        file_path: Word文档路径
    
    Returns:
        解析后的文本内容
    """
    content = []
    try:
        doc = Document(file_path)
        for para in doc.paragraphs:
            if para.text:
                content.append(para.text)
    except Exception as e:
        raise Exception(f"解析Word文档失败: {str(e)}")
    
    return '\n'.join(content)


def parse_excel(file_path: str) -> str:
    """解析Excel表格
    
    Args:
        file_path: Excel文件路径
    
    Returns:
        解析后的文本内容
    """
    content = []
    try:
        wb = load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            content.append(f"=== 工作表: {sheet_name} ===")
            
            # 读取所有行
            for row in sheet.iter_rows(values_only=True):
                # 过滤空行
                if any(cell is not None for cell in row):
                    # 转换为字符串并连接
                    row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                    content.append(row_text)
            content.append('')  # 工作表之间空一行
    except Exception as e:
        raise Exception(f"解析Excel文件失败: {str(e)}")
    
    return '\n'.join(content)


def parse_image(file_path: str) -> dict:
    """解析图片文件
    
    Args:
        file_path: 图片文件路径
    
    Returns:
        图片信息字典
    """
    try:
        with Image.open(file_path) as img:
            img_info = {
                "format": img.format,
                "size": img.size,
                "mode": img.mode,
                "filename": os.path.basename(file_path)
            }
        return img_info
    except Exception as e:
        raise Exception(f"解析图片文件失败: {str(e)}")


def parse_file(file_path: str) -> dict:
    """通用文件解析函数
    
    Args:
        file_path: 文件路径
    
    Returns:
        解析结果字典
    """
    # 获取文件类型
    file_type = get_file_type(file_path)
    filename = os.path.basename(file_path)
    
    # 根据文件类型选择解析器
    if file_type == 'application/pdf':
        content = parse_pdf(file_path)
        parsed_type = 'pdf'
    elif file_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document', 'application/msword']:
        content = parse_docx(file_path)
        parsed_type = 'docx'
    elif file_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']:
        content = parse_excel(file_path)
        parsed_type = 'excel'
    elif file_type.startswith('image/'):
        content = parse_image(file_path)
        parsed_type = 'image'
    else:
        # 文本文件直接读取
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            parsed_type = 'text'
        except UnicodeDecodeError:
            # 二进制文件
            content = f"无法解析的二进制文件: {file_type}"
            parsed_type = 'binary'
    
    return {
        "filename": filename,
        "file_path": file_path,
        "file_type": file_type,
        "parsed_type": parsed_type,
        "content": content
    }


def get_supported_file_types() -> list:
    """获取支持的文件类型列表
    
    Returns:
        支持的文件类型列表
    """
    return [
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "image/jpeg",
        "image/png",
        "image/gif",
        "image/bmp",
        "text/plain",
        "text/markdown",
        "text/html"
    ]
